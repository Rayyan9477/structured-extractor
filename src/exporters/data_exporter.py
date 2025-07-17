"""
Data Export Functionality for Medical Superbill Extraction

Implements comprehensive export capabilities for CSV, JSON, and Excel formats
with customizable field selection and formatting options.
"""

import csv
import json
import pandas as pd
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.core.config_manager import ConfigManager
from src.core.logger import get_logger
from src.core.data_schema import PatientData, ExtractionResults


class CSVExporter:
    """Exports patient data to CSV format."""
    
    def __init__(self, config: ConfigManager):
        """
        Initialize CSV exporter.
        
        Args:
            config: Configuration manager
        """
        self.config = config
        self.logger = get_logger(__name__)
        
        # Export configuration
        self.export_config = config.get("export", {}).get("csv", {})
        self.include_headers = self.export_config.get("include_headers", True)
        self.delimiter = self.export_config.get("delimiter", ",")
        self.encoding = self.export_config.get("encoding", "utf-8")
    
    def export_patients(self, patients: List[PatientData], output_path: str, flatten_codes: bool = True) -> None:
        """
        Export patient data to CSV.
        
        Args:
            patients: List of patient data to export
            output_path: Output file path
            flatten_codes: Whether to create separate rows for each CPT/ICD code
        """
        self.logger.info(f"Exporting {len(patients)} patients to CSV: {output_path}")
        
        try:
            if flatten_codes:
                rows = self._flatten_patient_data(patients)
            else:
                rows = self._create_summary_rows(patients)
            
            # Write to CSV
            with open(output_path, 'w', newline='', encoding=self.encoding) as csvfile:
                if rows:
                    writer = csv.DictWriter(csvfile, fieldnames=rows[0].keys(), delimiter=self.delimiter)
                    
                    if self.include_headers:
                        writer.writeheader()
                    
                    writer.writerows(rows)
            
            self.logger.info(f"Successfully exported to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to export CSV: {e}")
            raise
    
    def _flatten_patient_data(self, patients: List[PatientData]) -> List[Dict[str, Any]]:
        """Create flattened rows with separate entries for each medical code."""
        rows = []
        
        for patient in patients:
            base_row = self._create_base_patient_row(patient)
            
            # Create rows for each CPT code
            if patient.cpt_codes:
                for cpt in patient.cpt_codes:
                    row = base_row.copy()
                    row.update({
                        'code_type': 'CPT',
                        'medical_code': cpt.code,
                        'code_description': cpt.description,
                        'code_confidence': cpt.confidence
                    })
                    rows.append(row)
            
            # Create rows for each ICD-10 code
            if patient.icd10_codes:
                for icd in patient.icd10_codes:
                    row = base_row.copy()
                    row.update({
                        'code_type': 'ICD10',
                        'medical_code': icd.code,
                        'code_description': icd.description,
                        'code_confidence': icd.confidence
                    })
                    rows.append(row)
            
            # If no codes, create one row with patient info
            if not patient.cpt_codes and not patient.icd10_codes:
                rows.append(base_row)
        
        return rows
    
    def _create_summary_rows(self, patients: List[PatientData]) -> List[Dict[str, Any]]:
        """Create summary rows with concatenated codes."""
        rows = []
        
        for patient in patients:
            row = self._create_base_patient_row(patient)
            
            # Concatenate CPT codes
            if patient.cpt_codes:
                cpt_codes = [cpt.code for cpt in patient.cpt_codes]
                row['cpt_codes'] = '; '.join(cpt_codes)
            else:
                row['cpt_codes'] = ''
            
            # Concatenate ICD-10 codes
            if patient.icd10_codes:
                icd_codes = [icd.code for icd in patient.icd10_codes]
                row['icd10_codes'] = '; '.join(icd_codes)
            else:
                row['icd10_codes'] = ''
            
            rows.append(row)
        
        return rows
    
    def _create_base_patient_row(self, patient: PatientData) -> Dict[str, Any]:
        """Create base row with common patient information."""
        row = {
            'patient_index': patient.patient_index,
            'first_name': patient.first_name or '',
            'last_name': patient.last_name or '',
            'date_of_birth': patient.date_of_birth.isoformat() if patient.date_of_birth else '',
            'patient_id': patient.patient_id or '',
            'extraction_confidence': patient.extraction_confidence,
            'phi_detected': patient.phi_detected,
            'is_anonymized': getattr(patient, 'is_anonymized', False)
        }
        
        # Add service information
        if patient.service_info:
            row.update({
                'service_date': patient.service_info.date_of_service.isoformat() if patient.service_info.date_of_service else '',
                'provider_name': patient.service_info.provider_name or '',
                'provider_npi': patient.service_info.provider_npi or '',
                'facility_name': patient.service_info.facility_name or ''
            })
        else:
            row.update({
                'service_date': '',
                'provider_name': '',
                'provider_npi': '',
                'facility_name': ''
            })
        
        # Add financial information
        if patient.financial_info:
            row.update({
                'total_charges': patient.financial_info.total_charges,
                'insurance_payment': patient.financial_info.insurance_payment,
                'patient_payment': patient.financial_info.patient_payment,
                'balance_due': patient.financial_info.balance_due
            })
        else:
            row.update({
                'total_charges': None,
                'insurance_payment': None,
                'patient_payment': None,
                'balance_due': None
            })
        
        # Add contact information
        row.update({
            'phone': patient.phone or '',
            'address': patient.address or ''
        })
        
        return row


class JSONExporter:
    """Exports patient data to JSON format."""
    
    def __init__(self, config: ConfigManager):
        """
        Initialize JSON exporter.
        
        Args:
            config: Configuration manager
        """
        self.config = config
        self.logger = get_logger(__name__)
        
        # Export configuration
        self.export_config = config.get("export", {}).get("json", {})
        self.indent = self.export_config.get("indent", 2)
        self.ensure_ascii = self.export_config.get("ensure_ascii", False)
    
    def export_patients(self, patients: List[PatientData], output_path: str) -> None:
        """
        Export patient data to JSON.
        
        Args:
            patients: List of patient data to export
            output_path: Output file path
        """
        self.logger.info(f"Exporting {len(patients)} patients to JSON: {output_path}")
        
        try:
            # Convert patients to JSON-serializable format
            json_data = {
                'export_timestamp': datetime.now().isoformat(),
                'total_patients': len(patients),
                'patients': [self._patient_to_dict(patient) for patient in patients]
            }
            
            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=self.indent, ensure_ascii=self.ensure_ascii, default=self._json_serializer)
            
            self.logger.info(f"Successfully exported to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to export JSON: {e}")
            raise
    
    def export_extraction_results(self, results: ExtractionResults, output_path: str) -> None:
        """
        Export complete extraction results to JSON.
        
        Args:
            results: Extraction results to export
            output_path: Output file path
        """
        self.logger.info(f"Exporting extraction results to JSON: {output_path}")
        
        try:
            # Convert to JSON-serializable format
            json_data = {
                'file_path': results.file_path,
                'extraction_timestamp': results.extraction_timestamp.isoformat(),
                'total_patients': results.total_patients,
                'extraction_confidence': results.extraction_confidence,
                'metadata': results.metadata,
                'patients': [self._patient_to_dict(patient) for patient in results.patients]
            }
            
            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=self.indent, ensure_ascii=self.ensure_ascii, default=self._json_serializer)
            
            self.logger.info(f"Successfully exported extraction results to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to export extraction results: {e}")
            raise
    
    def _patient_to_dict(self, patient: PatientData) -> Dict[str, Any]:
        """Convert patient data to dictionary."""
        return patient.model_dump()
    
    def _json_serializer(self, obj):
        """Custom JSON serializer for datetime objects."""
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


class ExcelExporter:
    """Exports patient data to Excel format with formatting."""
    
    def __init__(self, config: ConfigManager):
        """
        Initialize Excel exporter.
        
        Args:
            config: Configuration manager
        """
        self.config = config
        self.logger = get_logger(__name__)
        
        # Export configuration
        self.export_config = config.get("export", {}).get("excel", {})
        self.include_formatting = self.export_config.get("include_formatting", True)
        self.freeze_panes = self.export_config.get("freeze_panes", True)
    
    def export_patients(self, patients: List[PatientData], output_path: str) -> None:
        """
        Export patient data to Excel.
        
        Args:
            patients: List of patient data to export
            output_path: Output file path
        """
        self.logger.info(f"Exporting {len(patients)} patients to Excel: {output_path}")
        
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Create summary sheet
            self._create_summary_sheet(workbook, patients)
            
            # Create detailed sheets
            self._create_detailed_sheets(workbook, patients)
            
            # Save workbook
            workbook.save(output_path)
            
            self.logger.info(f"Successfully exported to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to export Excel: {e}")
            raise
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, patients: List[PatientData]) -> None:
        """Create summary sheet with patient overview."""
        # Remove default sheet and create summary
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        sheet = workbook.create_sheet('Summary', 0)
        
        # Create DataFrame for summary
        summary_data = []
        for patient in patients:
            row = {
                'Patient Index': patient.patient_index,
                'Name': f"{patient.first_name or ''} {patient.last_name or ''}".strip(),
                'Date of Birth': patient.date_of_birth,
                'Patient ID': patient.patient_id,
                'CPT Codes': len(patient.cpt_codes) if patient.cpt_codes else 0,
                'ICD-10 Codes': len(patient.icd10_codes) if patient.icd10_codes else 0,
                'Service Date': patient.service_info.date_of_service if patient.service_info else None,
                'Provider': patient.service_info.provider_name if patient.service_info else '',
                'Total Charges': patient.financial_info.total_charges if patient.financial_info else None,
                'Confidence': patient.extraction_confidence,
                'PHI Detected': 'Yes' if patient.phi_detected else 'No'
            }
            summary_data.append(row)
        
        df = pd.DataFrame(summary_data)
        
        # Write DataFrame to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)
        
        # Apply formatting if enabled
        if self.include_formatting:
            self._format_summary_sheet(sheet)
        
        # Freeze panes
        if self.freeze_panes:
            sheet.freeze_panes = 'A2'
    
    def _create_detailed_sheets(self, workbook: openpyxl.Workbook, patients: List[PatientData]) -> None:
        """Create detailed sheets for CPT and ICD codes."""
        # CPT Codes sheet
        cpt_sheet = workbook.create_sheet('CPT Codes')
        cpt_data = []
        
        for patient in patients:
            if patient.cpt_codes:
                for cpt in patient.cpt_codes:
                    cpt_data.append({
                        'Patient Index': patient.patient_index,
                        'Patient Name': f"{patient.first_name or ''} {patient.last_name or ''}".strip(),
                        'CPT Code': cpt.code,
                        'Description': cpt.description,
                        'Confidence': cpt.confidence
                    })
        
        if cpt_data:
            cpt_df = pd.DataFrame(cpt_data)
            for r in dataframe_to_rows(cpt_df, index=False, header=True):
                cpt_sheet.append(r)
        
        # ICD-10 Codes sheet
        icd_sheet = workbook.create_sheet('ICD-10 Codes')
        icd_data = []
        
        for patient in patients:
            if patient.icd10_codes:
                for icd in patient.icd10_codes:
                    icd_data.append({
                        'Patient Index': patient.patient_index,
                        'Patient Name': f"{patient.first_name or ''} {patient.last_name or ''}".strip(),
                        'ICD-10 Code': icd.code,
                        'Description': icd.description,
                        'Confidence': icd.confidence
                    })
        
        if icd_data:
            icd_df = pd.DataFrame(icd_data)
            for r in dataframe_to_rows(icd_df, index=False, header=True):
                icd_sheet.append(r)
        
        # Apply formatting
        if self.include_formatting:
            self._format_detail_sheet(cpt_sheet)
            self._format_detail_sheet(icd_sheet)
        
        # Freeze panes
        if self.freeze_panes:
            cpt_sheet.freeze_panes = 'A2'
            icd_sheet.freeze_panes = 'A2'
    
    def _format_summary_sheet(self, sheet) -> None:
        """Apply formatting to summary sheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply header formatting
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_detail_sheet(self, sheet) -> None:
        """Apply formatting to detail sheets."""
        self._format_summary_sheet(sheet)  # Use same formatting


class DataExporter:
    """Main data exporter coordinating all export formats."""
    
    def __init__(self, config: ConfigManager):
        """
        Initialize data exporter.
        
        Args:
            config: Configuration manager
        """
        self.config = config
        self.logger = get_logger(__name__)
        
        # Initialize exporters
        self.csv_exporter = CSVExporter(config)
        self.json_exporter = JSONExporter(config)
        self.excel_exporter = ExcelExporter(config)
    
    def export_patients(
        self, 
        patients: List[PatientData], 
        output_path: str, 
        format_type: str = 'csv',
        **kwargs
    ) -> None:
        """
        Export patient data in specified format.
        
        Args:
            patients: List of patient data to export
            output_path: Output file path
            format_type: Export format ('csv', 'json', 'excel')
            **kwargs: Additional format-specific arguments
        """
        self.logger.info(f"Exporting {len(patients)} patients to {format_type}: {output_path}")
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        if format_type.lower() == 'csv':
            self.csv_exporter.export_patients(patients, output_path, **kwargs)
        elif format_type.lower() == 'json':
            self.json_exporter.export_patients(patients, output_path, **kwargs)
        elif format_type.lower() in ['excel', 'xlsx']:
            self.excel_exporter.export_patients(patients, output_path, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")
    
    def export_extraction_results(
        self, 
        results: ExtractionResults, 
        output_path: str, 
        format_type: str = 'json'
    ) -> None:
        """
        Export complete extraction results.
        
        Args:
            results: Extraction results to export
            output_path: Output file path
            format_type: Export format
        """
        if format_type.lower() == 'json':
            self.json_exporter.export_extraction_results(results, output_path)
        else:
            # For other formats, export just the patient data
            self.export_patients(results.patients, output_path, format_type)
    
    def export_batch(
        self, 
        results_list: List[ExtractionResults], 
        output_dir: str, 
        format_type: str = 'csv'
    ) -> None:
        """
        Export multiple extraction results.
        
        Args:
            results_list: List of extraction results
            output_dir: Output directory
            format_type: Export format
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        for i, results in enumerate(results_list):
            filename = f"extraction_results_{i+1:03d}.{format_type}"
            file_path = output_path / filename
            
            self.export_extraction_results(results, str(file_path), format_type)
